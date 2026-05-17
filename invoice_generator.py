import openpyxl
from fpdf import FPDF
import datetime
import os

def generate_invoice(client_name, service_desc, amount, invoice_number):
    """
    Erstellt ein PDF-Dokument für eine einzelne Rechnung.
    """
    pdf = FPDF()
    pdf.add_page()
    
    # Überschrift
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, txt="RECHNUNG / INVOICE", ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.ln(10)
    
    # Rechnungsdaten
    pdf.cell(0, 10, txt=f"Datum: {datetime.date.today()}", ln=True, align='R')
    pdf.cell(0, 10, txt=f"Rechnungsnummer: {invoice_number}", ln=True, align='L')
    pdf.ln(10)
        
    # Kunden- und Leistungsdaten
    pdf.cell(0, 10, txt=f"Kunde: {client_name}", ln=True, align='L')
    pdf.cell(0, 10, txt=f"Leistung: {service_desc}", ln=True, align='L')
    pdf.cell(0, 10, txt=f"Betrag: {amount} EUR", ln=True, align='L')
    
    # Fußzeile
    pdf.ln(20)
    pdf.set_font("Arial", 'I', 10)
    pdf.cell(0, 10, txt="Vielen Dank fuer Ihren Auftrag!", ln=True, align='C')
    
    # Ordner für Rechnungen erstellen, falls er nicht existiert
    if not os.path.exists("rechnungen_pdf"):
        os.makedirs("rechnungen_pdf")
        
    # PDF speichern
    filename = f"rechnungen_pdf/Rechnung_{invoice_number}.pdf"
    pdf.output(filename)
    print(f"[OK] Erstellt: {filename}")

def create_dummy_excel(filename="kundendaten.xlsx"):
    """
    Erstellt eine Beispiel-Excel-Tabelle zum Testen.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rechnungen"
    
    # Spaltenüberschriften
    ws.append(["Rechnungsnummer", "Kunde", "Leistung", "Betrag"])
    
    # Beispieldaten (Zahlen)
    ws.append(["2026-100", "Tech GmbH", "Softwareentwicklung", 4500.00])
    ws.append(["2026-101", "Baeckerei Meier", "Website Update", 350.50])
    ws.append(["2026-102", "StartUp UG", "Server Hosting", 120.00])
    
    wb.save(filename)
    print(f"--> Beispiel-Excel-Datei '{filename}' wurde automatisch generiert.")

def process_excel(file_path):
    """
    Liest die Excel-Tabelle Zeile für Zeile und generiert die PDFs.
    """
    print(f"Lese Daten aus: {file_path} ...")
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        count = 0
        # iter_rows liest die Zeilen ab der zweiten Zeile (min_row=2), ignoriert die Überschriften
        for row in sheet.iter_rows(min_row=2, values_only=True):
            invoice_number, client_name, service_desc, amount = row
            
            # Prüfen, ob die Zeile nicht leer ist
            if invoice_number and client_name:
                generate_invoice(str(client_name), str(service_desc), str(amount), str(invoice_number))
                count += 1
                
        print(f"=== FERTIG: {count} PDF-Rechnungen wurden aus der Excel-Tabelle generiert! ===")
    except Exception as e:
        print(f"FEHLER beim Lesen der Excel-Datei: {e}")

if __name__ == "__main__":
    excel_file = "kundendaten.xlsx"
    
    # 1. Erstelle eine Beispiel-Tabelle, falls keine existiert
    if not os.path.exists(excel_file):
        create_dummy_excel(excel_file)
        
    # 2. Wandle die Excel-Daten in PDFs um
    process_excel(excel_file)
